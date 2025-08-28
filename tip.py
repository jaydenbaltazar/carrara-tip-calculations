# main.py
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, time

# --- Allocation Table (Hardcoded) ---
ALLOCATION_TABLE = {
    'Lunch': {
        'Busser': 0.225, 'Barrista': 0.125, 'Kitchen': 0.1, 'Case': 0.15,
        'Register': 0.15, 'Lead': 0.25, 'Hostess': 0, 'Runner': 0
    },
    'Dinner_General': {
        'Busser': 0.2, 'Barrista': 0.13, 'Kitchen': 0.12, 'Case': .12,
        'Register': 0.18, 'Lead': 0.25, 'Hostess': .09, 'Runner': 0
    },
    'Dinner_Servers': {
        'Busser': 0.25, 'Barrista': 0.15, 'Kitchen': 0.15, 'Case': 0,
        'Register': 0.2, 'Lead': 0.25, 'Hostess': 0, 'Runner': 0
    },
    'Servers_to_Pool_Rate': 0.3
}

# --- Salary Employee Configuration ---
SALARY_EMPLOYEES = {
    'Jesus Elizondo': {
        'role': 'Kitchen',
        'weekly_hours': 80,
        'lunch_hours': 40,  # Split evenly between lunch and dinner
        'dinner_hours': 40
    }
}


def create_final_payroll_report(csv_file_path, xlsx_file_path, tips_csv_path=None):
    """
    Reads raw payroll data from a CSV, processes it based on complex role and
    time-based rules, and generates a formatted Excel summary report of hours worked.
    Now includes salary employees who aren't in the CSV file.
    """
    try:
        # --- 1. Read and Clean Raw Data ---
        df = pd.read_csv(csv_file_path)
        df.columns = df.columns.str.strip()
        df['Team Member'] = df['First'].str.strip() + ' ' + df['Last'].str.strip()
        df['Role'] = df['Role'].fillna('No Role')
        role_map = {
            'Dishwasher': 'Kitchen', 'Prep Cook': 'Kitchen', 'Pasta': 'Kitchen',
            'Salad': 'Kitchen', 'Grill': 'Kitchen', 'Shift Leader': 'Lead',
            'Host/Hostess': 'Hostess'
        }
        df['Role'] = df['Role'].str.strip().replace(role_map)
        split_roles = ['Busser', 'Barrista', 'Case', 'Register', 'Lead', 'Runner']
        df['In Time'] = pd.to_datetime(df['In Time'].str.strip(), format='%I:%M%p').dt.time
        df['Out Time'] = pd.to_datetime(df['Out Time'].str.strip(), format='%I:%M%p').dt.time
        
        # --- 2. Process Each Shift (The Core Logic) ---
        processed_data = []
        dinner_start_time = time(17, 0)  # 5:00 PM

        for index, row in df.iterrows():
            employee_name = row['Team Member']
            role = row['Role']
            total_hours = row['Regular hours']
            
            if role in split_roles:
                start_dt = datetime.combine(datetime.today(), row['In Time'])
                end_dt = datetime.combine(datetime.today(), row['Out Time'])
                if end_dt < start_dt: end_dt += pd.Timedelta(days=1)
                dinner_start_dt = datetime.combine(start_dt.date(), dinner_start_time)

                lunch_hours, dinner_hours = 0, 0
                if end_dt <= dinner_start_dt: lunch_hours = total_hours
                elif start_dt >= dinner_start_dt: dinner_hours = total_hours
                else:
                    lunch_hours = (dinner_start_dt - start_dt).total_seconds() / 3600
                    dinner_hours = (end_dt - dinner_start_dt).total_seconds() / 3600

                if lunch_hours > 0.001: processed_data.append([employee_name, f"{role}_Lunch", lunch_hours])
                if dinner_hours > 0.001: processed_data.append([employee_name, f"{role}_Dinner", dinner_hours])
            elif total_hours > 0:
                processed_data.append([employee_name, role, total_hours])

        # --- 2.5. Add Salary Employees ---
        print("Adding salary employees to payroll...")
        for employee_name, employee_config in SALARY_EMPLOYEES.items():
            role = employee_config['role']
            lunch_hours = employee_config['lunch_hours']
            dinner_hours = employee_config['dinner_hours']
            
            print(f"Adding {employee_name}: {lunch_hours} lunch hours + {dinner_hours} dinner hours = {lunch_hours + dinner_hours} total hours in {role}")

            # For Kitchen role, we don't split by lunch/dinner in the same way as other roles
            # Kitchen workers get their full hours counted as "Kitchen" (no _Lunch or _Dinner suffix)
            if role == 'Kitchen':
                total_kitchen_hours = lunch_hours + dinner_hours
                processed_data.append([employee_name, role, total_kitchen_hours])
            else:
                # For other roles that might be added in the future, split them
                if role in split_roles:
                    if lunch_hours > 0:
                        processed_data.append([employee_name, f"{role}_Lunch", lunch_hours])
                    if dinner_hours > 0:
                        processed_data.append([employee_name, f"{role}_Dinner", dinner_hours])
                else:
                    # For roles that don't split
                    total_hours = lunch_hours + dinner_hours
                    processed_data.append([employee_name, role, total_hours])

        # --- 3. Aggregate Data and Build Excel ---
        summary_df = pd.DataFrame(processed_data, columns=['Team Member', 'Role_Shift', 'Hours'])
        final_summary = summary_df.pivot_table(index='Team Member', columns='Role_Shift', values='Hours', aggfunc='sum').fillna(0)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Payroll Summary"

        # --- 4. Create Headers and Styles ---
        header_font = Font(bold=True, color="FFFFFF")
        centered_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        bold_font = Font(bold=True)
        main_header_fill = PatternFill(start_color="4472C4", fill_type="solid")
        sub_header_fill = PatternFill(start_color="D9D9D9", fill_type="solid")
        single_header_fill = PatternFill(start_color="808080", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        header_structure = {
            'Server': None, 'Busser': ['Lunch', 'Dinner'], 'Barrista': ['Lunch', 'Dinner'], 'Kitchen': None,
            'Case': ['Lunch', 'Dinner'], 'Register': ['Lunch', 'Dinner'], 'Training': None,
            'Lead': ['Lunch', 'Dinner'], 'Hostess': None, 'Runner': ['Lunch', 'Dinner'],
             'No Role': None
        }
        
        tm_cell = ws.cell(row=1, column=1, value='Team Member')
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        
        col_idx = 2
        for main_header, sub_headers in header_structure.items():
            if sub_headers:
                main_cell = ws.cell(row=1, column=col_idx, value=main_header)
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=1, end_column=col_idx + 1)
                ws.cell(row=2, column=col_idx, value=sub_headers[0])
                ws.cell(row=2, column=col_idx + 1, value=sub_headers[1])
                col_idx += 2
            else:
                main_cell = ws.cell(row=1, column=col_idx, value=main_header)
                ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
                col_idx += 1
        
        total_cell = ws.cell(row=1, column=col_idx, value='Total Hours')
        ws.merge_cells(start_row=1, start_column=col_idx, end_row=2, end_column=col_idx)
        
        # Apply styles to headers
        for r in ws.iter_rows(min_row=1, max_row=2, max_col=col_idx):
            for cell in r:
                cell.border = thin_border
                cell.alignment = centered_alignment
                if cell.row == 1: cell.font = header_font
                if cell.row == 2 and cell.value: cell.font = bold_font
                if cell.column == 1 or not header_structure.get(ws.cell(1, cell.column if cell.row==2 else cell.column).value):
                     cell.fill = single_header_fill
                elif cell.row == 1: cell.fill = main_header_fill
                else: cell.fill = sub_header_fill

        # --- 5. Write Data to Excel ---
        final_column_order = [f"{main}_{sub}" if sub else main for main, subs in header_structure.items() for sub in (subs or [''])]
        final_summary = final_summary.reindex(columns=final_column_order, fill_value=0)
        final_summary['Total Hours'] = final_summary.sum(axis=1)

        for r_idx, (index, row_data) in enumerate(final_summary.iterrows(), 3):
            ws.cell(row=r_idx, column=1, value=index)
            for c_idx, col_name in enumerate(final_summary.columns, 2):
                cell = ws.cell(row=r_idx, column=c_idx, value=row_data[col_name])
                cell.number_format = '0.00'
                cell.border = thin_border
                
                # Highlight salary employees with different background color
                if index in SALARY_EMPLOYEES:
                    cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")  # Light blue for salary employees

        # --- 6. Add Role Totals Row (RIGHT AFTER EMPLOYEE DATA) ---
        totals_row = len(final_summary) + 3  # Row immediately after employee data
        
        # Add "ROLE TOTALS" label
        totals_label_cell = ws.cell(row=totals_row, column=1, value='ROLE TOTALS')
        totals_label_cell.font = Font(bold=True, color="FFFFFF")
        totals_label_cell.fill = PatternFill(start_color="4472C4", fill_type="solid")
        totals_label_cell.alignment = centered_alignment
        totals_label_cell.border = thin_border
        
        # Calculate and add totals for each column
        for c_idx, col_name in enumerate(final_summary.columns, 2):
            if col_name != 'Total Hours':  # Skip the total hours column for now
                total_value = final_summary[col_name].sum()
                total_cell = ws.cell(row=totals_row, column=c_idx, value=total_value)
                total_cell.font = Font(bold=True)
                total_cell.number_format = '0.00'
                total_cell.border = thin_border
                total_cell.fill = PatternFill(start_color="E7E6E6", fill_type="solid")
        
        # Add grand total of all hours
        grand_total = final_summary.drop('Total Hours', axis=1).sum().sum()
        grand_total_cell = ws.cell(row=totals_row, column=len(final_summary.columns) + 1, value=grand_total)
        grand_total_cell.font = Font(bold=True)
        grand_total_cell.number_format = '0.00'
        grand_total_cell.border = thin_border
        grand_total_cell.fill = PatternFill(start_color="E7E6E6", fill_type="solid")

        # --- 7. Add space row after Role Totals ---
        space_row = totals_row + 1  # This row will be left empty for spacing
        
        # --- 7.5 Grab pre-calculated tips from file 
        if tips_csv_path:
            df = pd.read_csv(tips_csv_path, skiprows=6, header=None)

            df = df[[8, 15]]
            df.columns = ['Server', 'Tip']
            df.dropna(subset=['Server'], inplace=True)
            df = df[df['Server'] != 'Server']

            df['Tip'] = pd.to_numeric(df['Tip'], errors='coerce')
            df.dropna(subset=['Tip'], inplace=True)

            servers_tips = {}

            for index, row in df.iterrows():
                employee_name = row['Server']
                tip_value = row['Tip']

                # Fix: Split by just the comma and strip whitespace from parts
                name_parts = [part.strip() for part in employee_name.split(',')]
                
                # Check if the split was successful
                if len(name_parts) == 2:
                    employee_name = f"{name_parts[1]} {name_parts[0]}"
                
                if tip_value > 0:
                    servers_tips[employee_name] = {'tip': tip_value}

            print(servers_tips)

        # --- 8. Calculate Tips (if tips file provided) ---
        if tips_csv_path:
            try:
                # Read tips data
                tips_df = pd.read_csv(tips_csv_path)
                tips_df.columns = tips_df.columns.str.strip()
                
                # Find the row with "Total Allocated General Pool" to get lunch tips
                lunch_tips_total = 0
                dinner_tips_total = 0
                server_contribution_total = 0
                server_cash_cc_tips = 0
                
                # Look for the "Total Allocated General Pool" row
                for index, row in tips_df.iterrows():
                    if pd.notna(row.iloc[1]) and 'Total Allocated General Pool' in str(row.iloc[1]):
                        if pd.notna(row.iloc[2]):
                            lunch_tips_total = float(row.iloc[2])
                        if pd.notna(row.iloc[3]):
                            dinner_tips_total = float(row.iloc[3])
                        break
                
                # Look for the "Server Contribution to General Pool" row
                for index, row in tips_df.iterrows():
                    if pd.notna(row.iloc[1]) and 'Server Contribution to General Pool' in str(row.iloc[1]):
                        if pd.notna(row.iloc[3]):
                            server_contribution_total = float(row.iloc[3])
                        break
                
                # Look for the "Less Server Cash & CC Tips" row
                for index, row in tips_df.iterrows():
                    if pd.notna(row.iloc[1]) and 'Less Server Cash & CC Tips' in str(row.iloc[1]):
                        if pd.notna(row.iloc[3]):
                            server_cash_cc_tips = float(row.iloc[3])
                        break
                
                print(f"Found tips - Lunch: ${lunch_tips_total:.2f}, Dinner: ${dinner_tips_total:.2f}, Server Contribution: ${server_contribution_total:.2f}, Server Cash & CC: ${server_cash_cc_tips:.2f}")
                
                # --- LUNCH TIPS SECTION ---
                lunch_tips_row = space_row + 1
                
                # Add "LUNCH TIPS" label
                lunch_tips_label_cell = ws.cell(row=lunch_tips_row, column=1, value='LUNCH TIPS')
                lunch_tips_label_cell.font = Font(bold=True, color="FFFFFF")
                lunch_tips_label_cell.fill = PatternFill(start_color="70AD47", fill_type="solid")
                lunch_tips_label_cell.alignment = centered_alignment
                lunch_tips_label_cell.border = thin_border
                
                # Calculate tips for each lunch role
                lunch_tip_amounts = {}
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    role_tip_amount = 0
                    
                    if col_name != 'Total Hours':
                        if final_summary[col_name].sum() > 0:
                            if '_Lunch' in col_name:
                                role_name = col_name.replace('_Lunch', '')
                                if role_name in ALLOCATION_TABLE['Lunch']:
                                    allocation_rate = ALLOCATION_TABLE['Lunch'][role_name]
                                    role_tip_amount = lunch_tips_total * allocation_rate
                            elif col_name == 'Kitchen':
                                if 'Kitchen' in ALLOCATION_TABLE['Lunch']:
                                    allocation_rate = ALLOCATION_TABLE['Lunch']['Kitchen']
                                    role_tip_amount = lunch_tips_total * allocation_rate
                            elif col_name in ALLOCATION_TABLE['Lunch']:
                                allocation_rate = ALLOCATION_TABLE['Lunch'][col_name]
                                role_tip_amount = lunch_tips_total * allocation_rate
                        
                        lunch_tip_amounts[c_idx] = role_tip_amount
                        
                        # Add tip amount to Excel
                        tip_cell = ws.cell(row=lunch_tips_row, column=c_idx, value=role_tip_amount)
                        tip_cell.font = Font(bold=True)
                        tip_cell.number_format = '$0.00'
                        tip_cell.border = thin_border
                        tip_cell.fill = PatternFill(start_color="D5E8D4", fill_type="solid")
                
                # Add total lunch tips
                lunch_tip_total = sum(lunch_tip_amounts.values())
                total_tip_cell = ws.cell(row=lunch_tips_row, column=len(final_summary.columns) + 1, value=lunch_tip_total)
                total_tip_cell.font = Font(bold=True)
                total_tip_cell.number_format = '$0.00'
                total_tip_cell.border = thin_border
                total_tip_cell.fill = PatternFill(start_color="D5E8D4", fill_type="solid")
                
                # --- DINNER TIPS GENERAL SECTION ---
                dinner_tips_row = lunch_tips_row + 1
                
                # Add "DINNER TIPS GENERAL" label
                dinner_tips_label_cell = ws.cell(row=dinner_tips_row, column=1, value='DINNER TIPS GENERAL')
                dinner_tips_label_cell.font = Font(bold=True, color="FFFFFF")
                dinner_tips_label_cell.fill = PatternFill(start_color="FF6B35", fill_type="solid")
                dinner_tips_label_cell.alignment = centered_alignment
                dinner_tips_label_cell.border = thin_border
                
                # Calculate dinner tips for each role
                dinner_tip_amounts = {}
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    role_tip_amount = 0
                    
                    if col_name != 'Total Hours':
                        if final_summary[col_name].sum() > 0:
                            if '_Dinner' in col_name:
                                role_name = col_name.replace('_Dinner', '')
                                if role_name in ALLOCATION_TABLE['Dinner_General']:
                                    allocation_rate = ALLOCATION_TABLE['Dinner_General'][role_name]
                                    role_tip_amount = dinner_tips_total * allocation_rate
                            elif col_name == 'Kitchen':
                                if 'Kitchen' in ALLOCATION_TABLE['Dinner_General']:
                                    allocation_rate = ALLOCATION_TABLE['Dinner_General']['Kitchen']
                                    role_tip_amount = dinner_tips_total * allocation_rate
                            elif col_name in ALLOCATION_TABLE['Dinner_General']:
                                allocation_rate = ALLOCATION_TABLE['Dinner_General'][col_name]
                                role_tip_amount = dinner_tips_total * allocation_rate
                        
                        dinner_tip_amounts[c_idx] = role_tip_amount
                        
                        # Add tip amount to Excel
                        tip_cell = ws.cell(row=dinner_tips_row, column=c_idx, value=role_tip_amount)
                        tip_cell.font = Font(bold=True)
                        tip_cell.number_format = '$0.00'
                        tip_cell.border = thin_border
                        tip_cell.fill = PatternFill(start_color="FFE5DB", fill_type="solid")
                
                # Add total dinner tips
                dinner_tip_total = sum(dinner_tip_amounts.values())
                total_dinner_tip_cell = ws.cell(row=dinner_tips_row, column=len(final_summary.columns) + 1, value=dinner_tip_total)
                total_dinner_tip_cell.font = Font(bold=True)
                total_dinner_tip_cell.number_format = '$0.00'
                total_dinner_tip_cell.border = thin_border
                total_dinner_tip_cell.fill = PatternFill(start_color="FFE5DB", fill_type="solid")
                
                # --- DINNER TIPS SERVERS SECTION ---
                server_tips_row = dinner_tips_row + 1
                
                # Add "DINNER TIPS SERVERS" label
                server_tips_label_cell = ws.cell(row=server_tips_row, column=1, value='DINNER TIPS SERVERS')
                server_tips_label_cell.font = Font(bold=True, color="FFFFFF")
                server_tips_label_cell.fill = PatternFill(start_color="8E44AD", fill_type="solid")
                server_tips_label_cell.alignment = centered_alignment
                server_tips_label_cell.border = thin_border
                
                # Calculate server tips for each role
                server_tip_amounts = {}
                servers_to_pool_rate = ALLOCATION_TABLE['Servers_to_Pool_Rate']
                server_keep_amount = server_cash_cc_tips * (1 - servers_to_pool_rate)
                
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    role_tip_amount = 0
                    
                    if col_name != 'Total Hours':
                        if col_name == 'Server':
                            role_tip_amount = server_keep_amount
                        elif '_Dinner' in col_name:
                            role_name = col_name.replace('_Dinner', '')
                            if role_name in ALLOCATION_TABLE['Dinner_Servers']:
                                allocation_rate = ALLOCATION_TABLE['Dinner_Servers'][role_name]
                                role_tip_amount = server_contribution_total * allocation_rate
                        elif col_name == 'Kitchen':
                            if 'Kitchen' in ALLOCATION_TABLE['Dinner_Servers']:
                                allocation_rate = ALLOCATION_TABLE['Dinner_Servers']['Kitchen']
                                role_tip_amount = server_contribution_total * allocation_rate
                        elif col_name in ALLOCATION_TABLE['Dinner_Servers']:
                            allocation_rate = ALLOCATION_TABLE['Dinner_Servers'][col_name]
                            role_tip_amount = server_contribution_total * allocation_rate
                        
                        server_tip_amounts[c_idx] = role_tip_amount
                        
                        # Add tip amount to Excel
                        tip_cell = ws.cell(row=server_tips_row, column=c_idx, value=role_tip_amount)
                        tip_cell.font = Font(bold=True)
                        tip_cell.number_format = '$0.00'
                        tip_cell.border = thin_border
                        tip_cell.fill = PatternFill(start_color="E8DAEF", fill_type="solid")
                
                # Add total server tips
                server_tip_total = sum(server_tip_amounts.values())
                total_server_tip_cell = ws.cell(row=server_tips_row, column=len(final_summary.columns) + 1, value=server_tip_total)
                total_server_tip_cell.font = Font(bold=True)
                total_server_tip_cell.number_format = '$0.00'
                total_server_tip_cell.border = thin_border
                total_server_tip_cell.fill = PatternFill(start_color="E8DAEF", fill_type="solid")
                
                # --- GRAND TOTAL ROW (ALL TIPS COMBINED) ---
                grand_total_row = server_tips_row + 2  # Add a space row before grand total
                
                # Add "TOTAL" label
                grand_total_label = ws.cell(row=grand_total_row, column=1, value='TOTAL')
                grand_total_label.font = Font(bold=True, color="FFFFFF", size=12)
                grand_total_label.fill = PatternFill(start_color="2E2E2E", fill_type="solid")
                grand_total_label.alignment = centered_alignment
                grand_total_label.border = thin_border
                
                # Calculate grand totals for each column (sum of all three tip sections)
                total_role_tips = {}  # Store total tips per role for individual calculations
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    if col_name != 'Total Hours':
                        # Sum up tips from all three sections
                        lunch_amount = lunch_tip_amounts.get(c_idx, 0)
                        dinner_amount = dinner_tip_amounts.get(c_idx, 0)
                        server_amount = server_tip_amounts.get(c_idx, 0)
                        column_grand_total = lunch_amount + dinner_amount + server_amount
                        
                        # Store for individual calculations
                        total_role_tips[col_name] = column_grand_total
                        
                        # Add grand total to Excel
                        grand_total_cell = ws.cell(row=grand_total_row, column=c_idx, value=column_grand_total)
                        grand_total_cell.font = Font(bold=True, size=11)
                        grand_total_cell.number_format = '$0.00'
                        grand_total_cell.border = thin_border
                        grand_total_cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")
                
                # Add final grand total (sum of all tips)
                final_grand_total = lunch_tip_total + dinner_tip_total + server_tip_total
                final_total_cell = ws.cell(row=grand_total_row, column=len(final_summary.columns) + 1, value=final_grand_total)
                final_total_cell.font = Font(bold=True, size=11)
                final_total_cell.number_format = '$0.00'
                final_total_cell.border = thin_border
                final_total_cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")
                
                # --- INDIVIDUAL EMPLOYEE TIP CALCULATIONS ---
                individual_tips_start_row = grand_total_row + 3  # Add spacing before individual tips
                
                # Add "INDIVIDUAL EMPLOYEE TIPS" header
                individual_header = ws.cell(row=individual_tips_start_row, column=1, value='INDIVIDUAL EMPLOYEE TIPS')
                individual_header.font = Font(bold=True, color="FFFFFF", size=14)
                individual_header.fill = PatternFill(start_color="1F4E79", fill_type="solid")
                individual_header.alignment = centered_alignment
                individual_header.border = thin_border
                
                # Merge the header across all columns
                ws.merge_cells(start_row=individual_tips_start_row, start_column=1, end_row=individual_tips_start_row, end_column=len(final_summary.columns) + 1)
                
                # Add column headers for individual tips (same as original table)
                individual_header_row = individual_tips_start_row + 1
                
                # Team Member header
                tm_header = ws.cell(row=individual_header_row, column=1, value='Team Member')
                ws.merge_cells(start_row=individual_header_row, start_column=1, end_row=individual_header_row + 1, end_column=1)
                tm_header.font = header_font
                tm_header.fill = single_header_fill
                tm_header.alignment = centered_alignment
                tm_header.border = thin_border
                
                # Role headers
                col_idx = 2
                for main_header, sub_headers in header_structure.items():
                    if sub_headers:
                        main_cell = ws.cell(row=individual_header_row, column=col_idx, value=main_header)
                        ws.merge_cells(start_row=individual_header_row, start_column=col_idx, end_row=individual_header_row, end_column=col_idx + 1)
                        main_cell.font = header_font
                        main_cell.fill = main_header_fill
                        main_cell.alignment = centered_alignment
                        main_cell.border = thin_border
                        
                        lunch_cell = ws.cell(row=individual_header_row + 1, column=col_idx, value=sub_headers[0])
                        lunch_cell.font = bold_font
                        lunch_cell.fill = sub_header_fill
                        lunch_cell.alignment = centered_alignment
                        lunch_cell.border = thin_border
                        
                        dinner_cell = ws.cell(row=individual_header_row + 1, column=col_idx + 1, value=sub_headers[1])
                        dinner_cell.font = bold_font
                        dinner_cell.fill = sub_header_fill
                        dinner_cell.alignment = centered_alignment
                        dinner_cell.border = thin_border
                        col_idx += 2
                    else:
                        main_cell = ws.cell(row=individual_header_row, column=col_idx, value=main_header)
                        ws.merge_cells(start_row=individual_header_row, start_column=col_idx, end_row=individual_header_row + 1, end_column=col_idx)
                        main_cell.font = header_font
                        main_cell.fill = single_header_fill
                        main_cell.alignment = centered_alignment
                        main_cell.border = thin_border
                        col_idx += 1
                
                # Total Tips header
                total_tips_header = ws.cell(row=individual_header_row, column=col_idx, value='Total Tips')
                ws.merge_cells(start_row=individual_header_row, start_column=col_idx, end_row=individual_header_row + 1, end_column=col_idx)
                total_tips_header.font = header_font
                total_tips_header.fill = single_header_fill
                total_tips_header.alignment = centered_alignment
                total_tips_header.border = thin_border
                
                # Calculate individual employee tips
                individual_data_start_row = individual_header_row + 2
                
                # Get role totals from the ROLE TOTALS row we calculated earlier
                role_totals = {}
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    if col_name != 'Total Hours':
                        role_totals[col_name] = final_summary[col_name].sum()
                
                # Calculate tips for each employee
                for r_idx, (employee_name, employee_data) in enumerate(final_summary.iterrows()):
                    current_row = individual_data_start_row + r_idx
                    
                    # Add employee name
                    emp_name_cell = ws.cell(row=current_row, column=1, value=employee_name)
                    emp_name_cell.border = thin_border
                    emp_name_cell.alignment = Alignment(horizontal='left', vertical='center')
                    
                    # Highlight salary employees
                    if employee_name in SALARY_EMPLOYEES:
                        emp_name_cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")
                        emp_name_cell.font = Font(bold=True)
                    
                    employee_total_tips = 0
                    
                    # Calculate tips for each role
                    for c_idx, col_name in enumerate(final_summary.columns, 2):
                        if col_name != 'Total Hours':
                            employee_hours = employee_data[col_name]
                            role_total_hours = role_totals[col_name]
                            role_total_tips = total_role_tips.get(col_name, 0)
                            
                            # Calculate individual tip: (Total Role Tips) * (Employee Hours / Total Role Hours)
                            if role_total_hours > 0 and employee_hours > 0:
                                individual_tip = role_total_tips * (employee_hours / role_total_hours)
                            else:
                                individual_tip = 0
                                
                            
                            # --- FIX: Update the individual_tip for servers ---
                            if col_name == 'Server' and employee_name in servers_tips:
                                individual_tip = servers_tips[employee_name]['tip']
                                print(f"Corrected server tip for {employee_name} is {individual_tip}")
                            # --------------------------------------------------
                            
                            employee_total_tips += individual_tip
                                
                            
                            # Add to Excel
                            tip_cell = ws.cell(row=current_row, column=c_idx, value=individual_tip)
                            tip_cell.number_format = '$0.00'
                            tip_cell.border = thin_border
                            tip_cell.alignment = centered_alignment
                            
                            # Highlight salary employees and color coding for different tip amounts
                            if employee_name in SALARY_EMPLOYEES:
                                tip_cell.fill = PatternFill(start_color="E6F3FF", fill_type="solid")  # Light blue for salary employees
                            elif individual_tip > 0:
                                if individual_tip >= 50:
                                    tip_cell.fill = PatternFill(start_color="C6E0B4", fill_type="solid")  # Green for high tips
                                elif individual_tip >= 20:
                                    tip_cell.fill = PatternFill(start_color="FFE699", fill_type="solid")  # Yellow for medium tips
                                else:
                                    tip_cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")  # Light gray for low tips
                    
                    # Add total tips for employee
                    total_tip_cell = ws.cell(row=current_row, column=len(final_summary.columns) + 1, value=employee_total_tips)
                    total_tip_cell.number_format = '$0.00'
                    total_tip_cell.border = thin_border
                    total_tip_cell.alignment = centered_alignment
                    total_tip_cell.font = Font(bold=True)
                    
                    # Color code total based on amount, with special highlighting for salary employees
                    if employee_name in SALARY_EMPLOYEES:
                        total_tip_cell.fill = PatternFill(start_color="B3D9FF", fill_type="solid")  # Darker blue for salary employees
                    elif employee_total_tips >= 100:
                        total_tip_cell.fill = PatternFill(start_color="92D050", fill_type="solid")  # Bright green
                    elif employee_total_tips >= 50:
                        total_tip_cell.fill = PatternFill(start_color="C6E0B4", fill_type="solid")  # Light green
                    elif employee_total_tips >= 20:
                        total_tip_cell.fill = PatternFill(start_color="FFE699", fill_type="solid")  # Yellow
                    else:
                        total_tip_cell.fill = PatternFill(start_color="F2F2F2", fill_type="solid")  # Light gray
                
                # Add individual tips totals row
                individual_totals_row = individual_data_start_row + len(final_summary)
                
                # Add label
                totals_label = ws.cell(row=individual_totals_row, column=1, value='INDIVIDUAL TOTALS')
                totals_label.font = Font(bold=True, color="FFFFFF")
                totals_label.fill = PatternFill(start_color="1F4E79", fill_type="solid")
                totals_label.alignment = centered_alignment
                totals_label.border = thin_border
                
                # Calculate column totals for verification
                verification_total = 0
                for c_idx, col_name in enumerate(final_summary.columns, 2):
                    if col_name != 'Total Hours':
                        column_total = total_role_tips.get(col_name, 0)  # Should match the role total tips
                        verification_total += column_total
                        
                        total_cell = ws.cell(row=individual_totals_row, column=c_idx, value=column_total)
                        total_cell.font = Font(bold=True)
                        total_cell.number_format = '$0.00'
                        total_cell.border = thin_border
                        total_cell.fill = PatternFill(start_color="D9E2F3", fill_type="solid")
                        total_cell.alignment = centered_alignment
                
                # Add verification grand total
                verification_grand_total = ws.cell(row=individual_totals_row, column=len(final_summary.columns) + 1, value=verification_total)
                verification_grand_total.font = Font(bold=True)
                verification_grand_total.number_format = '$0.00'
                verification_grand_total.border = thin_border
                verification_grand_total.fill = PatternFill(start_color="D9E2F3", fill_type="solid")
                verification_grand_total.alignment = centered_alignment
                
                print(f"Lunch tips calculated: ${lunch_tip_total:.2f}")
                print(f"Dinner tips general calculated: ${dinner_tip_total:.2f}")
                print(f"Dinner tips servers calculated: ${server_tip_total:.2f}")
                print(f"GRAND TOTAL of all tips: ${final_grand_total:.2f}")
                
            except FileNotFoundError:
                print(f"Warning: Tips file '{tips_csv_path}' not found. Skipping tip calculations.")
            except Exception as e:
                print(f"Error processing tips file: {e}")

        # --- 9. Final Formatting and Save ---
        for col in ws.columns:
            ws.column_dimensions[get_column_letter(col[0].column)].width = 15
        ws.column_dimensions['A'].width = 25
        wb.save(xlsx_file_path)
        print(f"Successfully created payroll report: {xlsx_file_path}")
        
        # Print summary of salary employees added
        for emp_name, config in SALARY_EMPLOYEES.items():
            total_hours = config['lunch_hours'] + config['dinner_hours']
            print(f"Added salary employee: {emp_name} - {total_hours} hours in {config['role']}")
            
    except FileNotFoundError as e:
        print(f"Error: File not found - {e}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")

# --- Example Usage ---
#if __name__ == "__main__":
 #   # Step 1: Create the initial report with hours and tips
  #  hours_csv = "hours-and-wages-summary_2025-07-28_2025-08-10_all-locations.csv"
   # tips_csv = "tips-8-10.csv"  # Add your tips CSV file
    #payroll_report_xlsx = "tipFile_july-aug.xlsx"
    #create_final_payroll_report(hours_csv, payroll_report_xlsx, tips_csv)
    
    #using tips correctly & fix augustin name